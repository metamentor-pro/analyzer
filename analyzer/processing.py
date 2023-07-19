import openpyxl
import os
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import PatternFill


def check_cells_content(wb):
    sheet = wb.active

    cell_a1 = sheet['A1']
    cell_a2 = sheet['A2']
    cell_b1 = sheet['B1']

    if cell_a1.value is not None and cell_a1.value != "":
        return True
    if cell_a2.value is not None and cell_a2.value != "":
        return True
    if cell_b1.value is not None and cell_b1.value != "":
        return True

    return False


def get_data_range(sheet):
    start_row, start_col, end_row, end_col = None, None, None, None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                row_index = cell.row
                col_index = cell.column
                if start_row is None or row_index < start_row:
                    start_row = row_index
                if start_col is None or col_index < start_col:
                    start_col = col_index
                if end_row is None or row_index > end_row:
                    end_row = row_index
                if end_col is None or col_index > end_col:
                    end_col = col_index

    if start_row is not None and start_col is not None and end_row is not None and end_col is not None:
        start_col_name = openpyxl.utils.get_column_letter(start_col)
        end_col_name = openpyxl.utils.get_column_letter(end_col)
        data_range = f"{start_col_name}{start_row}:{end_col_name}{end_row}"
        return data_range, start_row, start_col, start_col_name, end_col_name
    else:
        return None


def move(wb):
    sheet = wb.active
    range = get_data_range(sheet)
    if range:
        sheet.move_range(range[0], rows=-range[1] + 1, cols=-range[2] + 1, translate=True)


def unmerge_cells(wb):
    sheet = wb.active
    range = get_data_range(sheet)

    if range is not None:
        check_range = f'{range[3]}{range[1]}:{range[4]}{range[1]}'
        is_merged = False
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.coord == check_range:
                is_merged = True
                break
        if is_merged:
            name = sheet.cell(row=range[1], column=range[2]).value

        mcr_coord_list = [mcr.coord for mcr in sheet.merged_cells.ranges]
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(mcr)
            if min_col == max_col:
                for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value
        if is_merged:
            sheet.delete_rows(range[1])


def process(path):
    save_path = path[:-5] + '_prepared.xlsx'
    name = os.path.basename(path)

    if not os.path.exists(save_path):
        wb = openpyxl.load_workbook(path, data_only=True)
        unmerge_cells(wb)
        if not check_cells_content(wb):
            move(wb)
        wb.save(save_path)
    return save_path, name


def unmerge_sheets(file_path):
    new_path = ""
    index = file_path.rfind("/")
    if index != -1:
        new_path = file_path[:index] + "/unmerged_{}/".format(os.path.basename(file_path))

    if not os.path.exists(new_path):
        check_path = file_path[:-5] + '_prepared.xlsx'
        if not os.path.exists(check_path):
            wb = openpyxl.load_workbook(file_path)
            if len(wb.sheetnames) > 1:
                os.makedirs(new_path)
                new_paths = list()
                for sheet_name in wb.sheetnames:
                    new_workbook = openpyxl.Workbook()
                    new_sheet = new_workbook.active
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        new_sheet.append(row)
                    new_workbook.save(new_path + sheet_name + '.xlsx')
                    new_paths.append(new_path + sheet_name + '.xlsx')
                return new_paths
            return [file_path]
        return [file_path]
    else:
        paths = list()
        for root, dirs, files in os.walk(new_path):
            for file in files:
                if not file.endswith('_prepared.xlsx'):
                    paths.append(os.path.join(root, file))
        return paths
