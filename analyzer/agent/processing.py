import os
import openpyxl
import pandas as pd
from openpyxl.utils.cell import range_boundaries
from typing import Union, List, Any


def check_cells_content(wb: openpyxl.workbook.workbook.Workbook) -> bool:
    sheet = wb.active
    cell_a1 = sheet['A1']
    cell_a2 = sheet['A2']
    cell_b1 = sheet['B1']
    if cell_a1.value is not None and cell_a1.value != "":
        return True
    if cell_a2.value is not None and cell_a2.value != "":
        return True
    if cell_b1.value is not None and cell_b1.value != "":
        return True
    return False


def get_data_range(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Any:
    start_row, start_col, end_row, end_col = None, None, None, None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                row_index = cell.row
                col_index = cell.column
                if start_row is None or row_index < start_row:
                    start_row = row_index
                if start_col is None or col_index < start_col:
                    start_col = col_index
                if end_row is None or row_index > end_row:
                    end_row = row_index
                if end_col is None or col_index > end_col:
                    end_col = col_index
    if start_row is not None and start_col is not None and end_row is not None and end_col is not None:
        start_col_name = openpyxl.utils.get_column_letter(start_col)
        end_col_name = openpyxl.utils.get_column_letter(end_col)
        data_range = f"{start_col_name}{start_row}:{end_col_name}{end_row}"
        return data_range, start_row, start_col, start_col_name, end_col_name
    else:
        return None


def move(wb: openpyxl.workbook.workbook.Workbook) -> None:
    sheet = wb.active
    table_range = get_data_range(sheet)
    if range:
        sheet.move_range(table_range[0], rows=-table_range[1] + 1, cols=-table_range[2] + 1, translate=True)


def unmerge_cells(wb: openpyxl.workbook.workbook.Workbook) -> None:
    sheet = wb.active
    table_range = get_data_range(sheet)
    if table_range is not None:
        check_range = f'{table_range[3]}{table_range[1]}:{table_range[4]}{table_range[1]}'
        is_merged = False
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.coord == check_range:
                is_merged = True
                break
        if is_merged:
            name = sheet.cell(row=table_range[1], column=table_range[2]).value
        mcr_coord_list = [mcr.coord for mcr in sheet.merged_cells.ranges]
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(mcr)
            if min_col == max_col:
                for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value
        if is_merged:
            sheet.delete_rows(table_range[1])


def process(path: str) -> tuple[str, str]:
    save_path = path[:-5] + '_prepared.xlsx'
    name = os.path.basename(path)
    if not os.path.exists(save_path):
        wb = openpyxl.load_workbook(path, data_only=True)
        unmerge_cells(wb)
        if not check_cells_content(wb):
            move(wb)
        wb.save(save_path)
    return save_path, name


def unmerge_sheets(file_path: str) -> List[str]:
    new_path = ""
    index = file_path.rfind("/")
    if index != -1:
        new_path = file_path[:index] + "/unmerged_{}/".format(os.path.basename(file_path))
    if not os.path.exists(new_path):
        check_path = file_path[:-5] + '_prepared.xlsx'
        if not os.path.exists(check_path):
            excel_file = pd.ExcelFile(file_path)
            if len(excel_file.sheet_names) > 1:
                os.makedirs(new_path)
                new_paths = list()
                for sheet_name in excel_file.sheet_names:
                    df = excel_file.parse(sheet_name)
                    df.to_excel(new_path + sheet_name + '.xlsx', index=False)
                    new_paths.append(new_path + sheet_name + '.xlsx')
                return new_paths
            return [file_path]
        return [file_path]
    else:
        paths = list()
        for root, dirs, files in os.walk(new_path):
            for file in files:
                if not file.endswith('_prepared.xlsx'):
                    paths.append(os.path.join(root, file))
        return paths
